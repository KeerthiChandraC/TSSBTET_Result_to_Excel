import openpyxl
from openpyxl.utils import get_column_letter
from time import sleep
import os
from bs4 import BeautifulSoup
thisdir = os.getcwd()


def coderintro():
    screen_time =4
    for i in range(0,screen_time):
                os.system('cls')
                print()
                print("********************************************************************************")
                print()
                print("Program for spliting cells and save in excel for C18 and C21 Results")
                print("********************************************************************************")
                print()
                print("********************************************************************************")
                print()
                print("developed by Keerthi Chandra C")
                print("Lecturer in ECE , GMRPW Karimnagar")
                print("for constructive criticism and suggestions reach me at keerthichand.c@gmail.com")
                print('''


disclaimer:
    *These are not actual Results
    *Use this at your own discretion
    *for actual Results please visit TS-SBTET Website
    *OUTPUTS from this program are not related to TS-SBTET
    *This program is developed only to help ease work but not to misguide or harm in 
        any possible ways and with no ill intent
    ''')
                print("******************************************************************************")
                if screen_time-i-1>0:
                        print(f'starting in {screen_time-i-2}s .......')
                sleep(1)
    input("press ENTER to start.....")
    os.system('cls')

def instructions_info():
        print('''
************************************************************************************

before staring ...
Save C18/C21 Regular Branch Results directly from browser as html fie
Paste your C18/C21 Regular Branch Results saved from SBTET portal in INPUT_HTML folder in html format

************************************************************************************''')
        input("press ENTER to continue......")
        os.system('cls')
        print()

def read(head,body,fname):


    hrows = head.find_all('tr')
    for tr in hrows:
        td = tr.find_all('th')
           
        headrow = [i.text for i in td]
    headrow = headrow[2:12]
    
    
    wb = openpyxl.Workbook()
    Marks_sheet = wb.active
    wb_old = openpyxl.Workbook()
    Marks_sheet_old = wb_old.active
    row_pad =2
    Marks_sheet.cell(row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    Marks_sheet.cell(row=row_pad, column=1).value = "SlNo"
    Marks_sheet.cell(row=row_pad, column=2).value = "PIN"
    Marks_sheet.cell(row=row_pad, column=3).value = "NAME"
    Marks_sheet_old.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    Marks_sheet_old.cell(row=4, column=1).value = "Sl No"
    Marks_sheet_old.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    Marks_sheet_old.cell(row=4, column=2).value = "PIN"
    Marks_sheet_old.cell(row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    Marks_sheet_old.cell(row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
    #Marks_sheet_old.cell(row=4, column=2+1+7*j).value = mid1

    col =4
    j=0
    headrow = [splith(a) for a in headrow ]
    for sub in headrow:
        Marks_sheet_old.merge_cells(start_row=3, start_column=2+1+7*j, end_row=3, end_column=2+7+7*j)
        Marks_sheet_old.cell(row=3, column=2+1+7*j).value = sub
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+1+7*j, end_row=5, end_column=2+1+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+2+7*j, end_row=5, end_column=2+2+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+3+7*j, end_row=5, end_column=2+3+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+4+7*j, end_row=5, end_column=2+4+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+6+7*j, end_row=5, end_column=2+6+7*j)
        Marks_sheet_old.cell(row=4, column=2+1+7*j).value = "MID1"
        Marks_sheet_old.cell(row=4, column=2+2+7*j).value = "MID2"
        Marks_sheet_old.cell(row=4, column=2+3+7*j).value = "EXT"
        Marks_sheet_old.cell(row=4, column=2+4+7*j).value = "TOTAL_EXT"
        Marks_sheet_old.cell(row=5, column=2+5+7*j).value = ""
        Marks_sheet_old.cell(row=4, column=2+5+7*j).value = "EXT_AVG"
        Marks_sheet_old.cell(row=4, column=2+6+7*j).value = "INT"
        Marks_sheet_old.cell(row=5, column=2+7+7*j).value = ""
        Marks_sheet_old.cell(row=4, column=2+7+7*j).value = "INT_AVG"
        
        j= j+1
        for stat in [f"{sub}_mid1",f"{sub}_mid2",f"{sub}_intr",f"{sub}_ext",f"{sub}_total",f"{sub}_grade",f"{sub}_status"]:
        
            Marks_sheet.cell(row=row_pad, column=col).value = stat
            #print(stat, end = ' ')
            col=col+1
            
        
    for stat in ["RUBRICS","CREDITS","TOTALMARKS","TOTALGRADE","SGPA","CGPA","RESULT"]:
                Marks_sheet.cell(row=row_pad, column=col).value = stat
                col=col+1

    
    rows = body.find_all('tr')    #print(rows)
    i=1
    for row in rows:
            cols = row.find_all('td')
            if len(cols)!=19:
                continue
            #print(cols[0].text)
            cellsa = [cell.text for cell in cols ]
            #print(len(cellsa))
            #print(cellsa)
            if '-'in cellsa[0]:
                PIN= cellsa[0]
                NAME= cellsa[1]
                R18EC101F= cellsa[2]
                R18EC102F= cellsa[3]
                R18EC103F= cellsa[4]
                R18EC104F= cellsa[5]
                R18EC105C= cellsa[6]
                R18EC106P= cellsa[7]
                R18EC107P= cellsa[8]
                R18EC108P= cellsa[9]
                R18EC109P= cellsa[10]
                R18EC110P= cellsa[11]
                RUBRICS= cellsa[12]
                CREDITS= cellsa[14]
                TOTALmARKS= cellsa[13]
                TOTALGRADE= cellsa[15]
                SGPA= cellsa[16]
                CGPA= cellsa[17]
                RESULT = cellsa[18]
                exfill(Marks_sheet,i+1,PIN,NAME,R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P,RUBRICS,CREDITS,TOTALGRADE,SGPA,CGPA,RESULT,TOTALmARKS)
                exfill_old(Marks_sheet_old,i,PIN, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P)
                
                i=i+1
    os.makedirs(f'OUTPUTS', exist_ok=True)
    wb.save(f'OUTPUTS/{fname}_Type1.xlsx')

    wb_old.save(f'OUTPUTS/{fname}_Type2.xlsx')

def readht(file):
    soup = BeautifulSoup(open(f'{thisdir}/INPUT_HTML/{file}'),'html.parser')
    tables = soup.find_all('table')
    for table in tables:
        heads = table.find_all('thead')
        body = table.find_all('tbody')
        for head in heads:
            rows = head.find_all('tr')
            for tr in rows:
                td = tr.find_all('th')
                   
                row = [i.text for i in td]
                if len(row)== 19 and "PIN" in row[0].upper():
                    read(head,body[0],file.split('.')[0])
        


def splitc(a):
    b = a.replace('\n','')
    b = b.replace(' ','')
    b = b.split('(')
    a = a.split('(')
    a = a[1]
    t = a.split("\n")[1]
    a = a.split(')')
    
    a = a[0]
    a = a.split("+")
    return a[0],a[1],a[2],a[3],t,b[0]
def splith(a):
    b = a.split('(')[0]
    b = b.replace('\n','')
    b = b.replace(' ','')
    
    return b
    

def exfill(Marks_sheet,i,PIN,NAME, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P,RUBRICS,CREDITS,TOTALGRADE,SGPA,CGPA,RESULT,TOTALmARKS):
    row_pad =1
    Marks_sheet.cell(row=i+row_pad, column=1).value = int(i-1)
    Marks_sheet.cell(row=i+row_pad, column=2).value = str(PIN)
    Marks_sheet.cell(row=i+row_pad, column=3).value = str(NAME)
    print(i-1, PIN,end=' ')
    col =4
    for sub in (R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
        #print(sub)
        mid1,mid2,intr,ext,total,grade = splitc(sub)
        #print(mid1,mid2,intr,ext,total,grade)
        try:
            mid1 =float(mid1)
        except:
            mid1 =0

        try:
            mid2 =float(mid2)
        except:
            mid2 =0
        try:
            intr =float(intr)
        except:
            intr =0
        try:
            ext =float(ext)
        except:
            ext =0
        total = mid1+mid2+intr+ext
        status = "PASS" if grade !='E' else "Fail"

        
        for stat in [mid1,mid2,intr,ext,total,grade,status]:
            
            Marks_sheet.cell(row=i+row_pad, column=col).value = stat
            print(stat, end = ' ')
            col=col+1
            
        
    for stat in [RUBRICS,CREDITS,TOTALmARKS,TOTALGRADE,SGPA,CGPA,RESULT]:
                try:
                    stat =float(stat)
                except:
                    stat =stat
        
                Marks_sheet.cell(row=i+row_pad, column=col).value = stat
                col=col+1
                
    print('')

def exfill_old(Marks_sheet_old,i,PIN, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
    
    Marks_sheet_old.cell(row=i+5, column=1).value = int(i)
    Marks_sheet_old.cell(row=i+5, column=2).value = str(PIN)
    #print(i, PIN,end=' ')
    pad =2
    j=0
    for x in (R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
        mid1,mid2,intr,ext,total,grade = splitc(x)
        try:
            mid1 =float(mid1)
        except:
            mid1 =0

        try:
            mid2 =float(mid2)
        except:
            mid2 =0
        try:
            intr =float(intr)
        except:
            intr =0
        try:
            ext =float(ext)
        except:
            ext =0
        EXTERN_EXAM = float(mid1)+float(mid2)+float(ext)
        INTERN_EXAM = float(intr)
        
        
        Marks_sheet_old.cell(row=i+5, column=pad+1+7*j).value = mid1
        Marks_sheet_old.cell(row=i+5, column=pad+2+7*j).value = mid2
        Marks_sheet_old.cell(row=i+5, column=pad+3+7*j).value = ext
        Marks_sheet_old.cell(row=i+5, column=pad+4+7*j).value = EXTERN_EXAM#get_column_letter(1)
        Marks_sheet_old.cell(row=i+5, column=pad+5+7*j).value = f'=IF({get_column_letter(pad+4+7*j)}{i+5}>={get_column_letter(pad+4+7*j+1)}$5,"Y","N")'
        Marks_sheet_old.cell(row=5, column=pad+5+7*j).value = f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+5+7*j-1)}6:{get_column_letter(pad+5+7*j-1)}{i+5}),0)'
        #print(f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+4+7*j-1)}5:{get_column_letter(pad+6+7*j-1)}{i+5},0))')
        #Marks_sheet_old.cell(row=i+4, column=pad+5+7*j).value =  EX
        Marks_sheet_old.cell(row=i+5, column=pad+6+7*j).value = INTERN_EXAM
        Marks_sheet_old.cell(row=i+5, column=pad+7+7*j).value = f'=IF({get_column_letter(pad+6+7*j)}{i+5}>={get_column_letter(pad+6+7*j+1)}$5,"Y","N")'
        Marks_sheet_old.cell(row=5, column=pad+7+7*j).value = f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+7+7*j-1)}6:{get_column_letter(pad+7+7*j-1)}{i+5}),0)'
        #Marks_sheet_old.cell(row=i+4, column=pad+7+7*j).value = INTX
        j=j+1
        #print(f'{mid1}, {mid2}, {intr}, {ext}', end=' ')

    #print('')
if __name__=='__main__':
    count = 0
    coderintro()
    instructions_info()
    for r, d, f in os.walk(f'{thisdir}/INPUT_HTML'):
        for file in f:
            if file.endswith(".html") or file.endswith(".htm"):
                           print(f"Converting {file}")
                           readht(file)
                           print(f"Conversion copleted......")
                           count +=1
    if count >0:
        print("All files are saved in Outputs folder")
        input("Press Enter to Exit")
    else:
        print("No html files found")
        input("Press Enter to Exit")
                
                
    
