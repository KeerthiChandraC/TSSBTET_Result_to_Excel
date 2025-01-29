import openpyxl
from openpyxl.utils import get_column_letter
from time import sleep
import os
import pandas as pd
import urllib3
urllib3.disable_warnings()
import requests
from openpyxl.styles import Alignment
from datetime import date
from json import loads
from collections import OrderedDict

c16s = "https://www.sbtet.telangana.gov.in/API/api/Results/GetC16SConsolidatedResult?Pin="
c18 = "https://www.sbtet.telangana.gov.in/API/api/Results/GetConsolidatedResults?Pin="
c21 = "https://www.sbtet.telangana.gov.in/API/api/Results/GetConsolidatedResults?Pin="
c16 = "https://www.sbtet.telangana.gov.in/API/api/Results/GetC16ConsolidatedResult?Pin="


TOP_ROW_tot = ['Scheme', 'StudentName', 'Pin', 'BranchCode']
TOP_ROW_SEM = ['Credits', 'SGPA', 'PASSED_TYPE', 'PASSED_STATUS']
TOP_ROW_gen = ['StudentName', 'Pin', 'BranchCode',
               'Scheme', 'CGPA', 'TOTAL_CreditsGained', 'Diploma_Status']

file_path = os.getcwd()

TODAY = date.today().strftime("%Y-%m-%d")

SEM_INSERT = {"1st_Year_Passed":16, 
              "2nd_Year_Passed":25,
              "3rd_Year_Passed":34}

SEM_CONT = [1,5]



thisdir = os.getcwd()

SLOW_LEARNERS_PERCENT  = 40
FAST_LEARNERS_PERCENT  = 75

INTERNAL_MAX_MARKS = 20
MID1_MAX_MARKS = 20
MID2_MAX_MARKS = 20
EXTERNAL_MAX_MARKS = 40
TOTAL_MAX_MARKS = INTERNAL_MAX_MARKS+MID1_MAX_MARKS+MID2_MAX_MARKS+EXTERNAL_MAX_MARKS 

RESULT_TABLE_INDEX = 2
RESULT_TABLE_LENGTH = 19
SEM6_RESULT_TABLE_LENGTH = 10

def coderintro():
    screen_time =4
    for i in range(0,screen_time):
                os.system('cls')
                print()
                print("********************************************************************************")
                print()
                print("Program for spliting cells and save in excel for C18,C21 and C24 Results with subject wise result analysis")
                print("Program also seperates slow learners(below 40%) and Fast Learners(above 75%)")
                print()
                print("********************************************************************************")
                print()
                print("********************************************************************************")
                print()
                print("developed by Keerthi Chandra C")
                print("Lecturer in ECE,")
                print("for constructive criticism and suggestions reach me at")
                print('''Phone: 9963459392
Email: keerthichand.c@gmail.com''')
                print('''


disclaimer:
    *These are not actual Results
    *Use this at your own discretion
    *for actual Results please visit TS-SBTET Website
    *OUTPUTS from this program are not related to TS-SBTET
    *This program is developed only to help ease work but not to misguide or harm in 
        any possible ways and with no ill intent
    ''')
                print("******************************************************************************")
                if screen_time-i-1>0:
                        print(f'starting in {screen_time-i-2}s .......')
                sleep(1)
    input("press ENTER to start.....")
    os.system('cls')

def instructions_info():
        print('''
************************************************************************************

before staring ...
Save C18/C21 Regular Branch Results directly from browser as html fie
Paste your C18/C21 Regular Branch Results saved from SBTET portal in INPUT_HTML folder in html format
for detailed instructions watch below video:
    https://youtu.be/-5EINkWqOg4

************************************************************************************''')
        input("press ENTER to continue......")
        os.system('cls')
        print()

def getpins(result_df):
    pins = []
   
    for pin in result_df['Pin']:
        if not(pin == None) and '-' in pin:
            pin = pin.upper()
            pin = pin.replace("\n","")
            pin = pin.replace(" ","")
            pins.append(pin)
            
    #for pin in pins:
        #print(pin)
    return pins

def conect(url):
    timeout = 5
    op = None
    while True:
        try:
            r = requests.get(url,verify = False, timeout=timeout)
            op = r.status_code
            if op == 200:
                pass
                # print("Done.")
                break
            else:
                # print("Status Code is not 200")
                # print("status Code", op)
                pass
        except:
            # print("Not Connected. No internet.")
            # print("trying again...")
            sleep(5)

    return r

def conect_with_print(url):
    timeout = 5
    op = None
    #print(url)
    while True:
        try:
            r = requests.get(url, verify = False, timeout=timeout)
            op = r.status_code
            #print(r)
            if op == 200:
                pass
                print("Done.")
                break
            else:
                print("Status Code is not 200")
                print("status Code", op)
                pass
        except Exception as e:
            #print(e)
            #print("status Code", op)
            print('''Unable to Connect to server.
Please check your internet connection.
if internet is okay.
Contact keerthichand.c@gmail.com''')
            print("trying again...")
            sleep(5)
    return r

def clean_df(df):
    if len(df.columns) == SEM6_RESULT_TABLE_LENGTH:
        head =  list(df)
        for i in range(0,RESULT_TABLE_LENGTH-SEM6_RESULT_TABLE_LENGTH):
            df.insert(2,f'{head[2]}_{i}',df[head[2]])
    if len(df.columns) > RESULT_TABLE_LENGTH:
        head =  list(df)
        clean_Head = list(OrderedDict.fromkeys([f"{x.split('-')[0]}-{x.split('-')[1][:3]}" if "-" in x else x for x in head]))
        df2 = df.iloc[:,:19]
        
        try:
            df2.columns = clean_Head
        except:
           clean_Head = ["Pin","Name","501","502","503","504","505","506","507","508","509","510",'Rubrics','Total','Credits','Total Grade Points','SGPA','CGPA','Result'] 
           df2.columns = clean_Head
        df = df2  
    return df       
def rsultAnaly(df,fname):
    df = df.map(lambda s: s.upper() if type(s) == str else s)
    #print(df)
    pas = df['Result'].value_counts()['PASS']
    total = df['Result'].count()
    fail =  total - pas
    data_dict = {f"{fname}":["Total Students:","No of Passed Students:","No of Failed Students:", "Pass Percentage:"],
            f"Result Analysis":[f"{total}",f"{pas}",f"{fail}", f"{round((pas/total)*100,2)}"]}                    
    data  = pd.DataFrame(data_dict)
    data.to_excel(f'OUTPUTS/{fname}_ResultAnalysis.xlsx')     

def PD_readht(file):
    tables= pd.read_html(f'{thisdir}/INPUT_HTML/{file}')
    fname = file.split('.')[0]
    result_df = tables[RESULT_TABLE_INDEX]
    result_df.to_excel(f'OUTPUTS/{fname}_RawResult.xlsx')
    if len(result_df.columns) != RESULT_TABLE_LENGTH:    
         result_df = clean_df(result_df)
    #print(result_df)
    result_df = result_df.dropna()
    result_df = result_df.reset_index(drop=True)
    rsultAnaly(result_df,fname)
    pins = getpins(result_df)
    #print(pins)
    get_cons(pins,fname)
    #print(result_df.head())
    pd_read(result_df,fname)
    
def pd_read(df,fname):

    headrow = list(df)
    print(headrow)
    headrow = headrow[2:12]
    
    
    wb = openpyxl.Workbook()
    Marks_sheet = wb.active
    wb_old = openpyxl.Workbook()
    Marks_sheet_old = wb_old.active
    wb_old2 = openpyxl.Workbook()
    Marks_sheet_old2 = wb_old2.active
    row_pad =2
    Marks_sheet.cell(row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    Marks_sheet.cell(row=row_pad, column=1).value = "SlNo"
    Marks_sheet.cell(row=row_pad, column=2).value = "PIN"
    Marks_sheet.cell(row=row_pad, column=3).value = "NAME"
    Marks_sheet_old.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    Marks_sheet_old.cell(row=4, column=1).value = "Sl No"
    Marks_sheet_old.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    Marks_sheet_old.cell(row=4, column=2).value = "PIN"
    Marks_sheet_old.cell(row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    Marks_sheet_old.cell(row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"

    Marks_sheet_old2.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    Marks_sheet_old2.cell(row=4, column=1).value = "Sl No"
    Marks_sheet_old2.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    Marks_sheet_old2.cell(row=4, column=2).value = "PIN"
    Marks_sheet_old2.cell(row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    Marks_sheet_old2.cell(row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
    #Marks_sheet_old.cell(row=4, column=2+1+7*j).value = mid1

    col =4
    j=0
    headrow = [splith(a) for a in headrow ]
    for sub in headrow:
        Marks_sheet_old.merge_cells(start_row=3, start_column=2+1+7*j, end_row=3, end_column=2+7+7*j)
        Marks_sheet_old.cell(row=3, column=2+1+7*j).value = sub
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+1+7*j, end_row=5, end_column=2+1+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+2+7*j, end_row=5, end_column=2+2+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+3+7*j, end_row=5, end_column=2+3+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+4+7*j, end_row=5, end_column=2+4+7*j)
        Marks_sheet_old.merge_cells(start_row=4, start_column=2+6+7*j, end_row=5, end_column=2+6+7*j)
        Marks_sheet_old.cell(row=4, column=2+1+7*j).value = "MID1"
        Marks_sheet_old.cell(row=4, column=2+2+7*j).value = "MID2"
        Marks_sheet_old.cell(row=4, column=2+3+7*j).value = "EXT"
        Marks_sheet_old.cell(row=4, column=2+4+7*j).value = "TOTAL_EXT"
        Marks_sheet_old.cell(row=5, column=2+5+7*j).value = ""
        Marks_sheet_old.cell(row=4, column=2+5+7*j).value = "EXT_AVG"
        Marks_sheet_old.cell(row=4, column=2+6+7*j).value = "INT"
        Marks_sheet_old.cell(row=5, column=2+7+7*j).value = ""
        Marks_sheet_old.cell(row=4, column=2+7+7*j).value = "INT_AVG"

        Marks_sheet_old2.merge_cells(start_row=3, start_column=2+1+7*j, end_row=3, end_column=2+7+7*j)
        Marks_sheet_old2.cell(row=3, column=2+1+7*j).value = sub
        Marks_sheet_old2.merge_cells(start_row=4, start_column=2+1+7*j, end_row=5, end_column=2+1+7*j)
        Marks_sheet_old2.merge_cells(start_row=4, start_column=2+2+7*j, end_row=5, end_column=2+2+7*j)
        Marks_sheet_old2.merge_cells(start_row=4, start_column=2+3+7*j, end_row=5, end_column=2+3+7*j)
        Marks_sheet_old2.merge_cells(start_row=4, start_column=2+4+7*j, end_row=5, end_column=2+4+7*j)
        Marks_sheet_old2.merge_cells(start_row=4, start_column=2+6+7*j, end_row=5, end_column=2+6+7*j)
        Marks_sheet_old2.cell(row=4, column=2+1+7*j).value = "MID1"
        Marks_sheet_old2.cell(row=4, column=2+2+7*j).value = "MID2"
        Marks_sheet_old2.cell(row=4, column=2+3+7*j).value = "INT"#"EXT"
        Marks_sheet_old2.cell(row=4, column=2+4+7*j).value = "TOTAL_INT"#"TOTAL_EXT"
        Marks_sheet_old2.cell(row=5, column=2+5+7*j).value = ""
        Marks_sheet_old2.cell(row=4, column=2+5+7*j).value = "INT_AVG"#"EXT_AVG"
        Marks_sheet_old2.cell(row=4, column=2+6+7*j).value = "EXT"#"INT"
        Marks_sheet_old2.cell(row=5, column=2+7+7*j).value = ""
        Marks_sheet_old2.cell(row=4, column=2+7+7*j).value = "EXT_AVG"#"INT_AVG"
        
        j= j+1
        for stat in [f"{sub}_mid1",f"{sub}_mid2",f"{sub}_intr",f"{sub}_ext",f"{sub}_total",f"{sub}_grade",f"{sub}_status"]:
        
            Marks_sheet.cell(row=row_pad, column=col).value = stat
            #print(stat, end = ' ')
            col=col+1
            
        
    for stat in ["RUBRICS","CREDITS","TOTALMARKS","TOTALGRADE","SGPA","CGPA","RESULT"]:
                Marks_sheet.cell(row=row_pad, column=col).value = stat
                col=col+1

    
    rows = df.values.tolist()    #print(rows)
    #print(rows)
    i=1
    for row in rows:
            cellsa = row
            if len(cellsa)!=RESULT_TABLE_LENGTH:
                continue
            #print(cols[0].text)
            #cellsa = 
            #print(len(cellsa))
            #print(cellsa)
            if '-'in cellsa[0]:
                PIN= cellsa[0]
                NAME= cellsa[1]
                R18EC101F= cellsa[2]
                R18EC102F= cellsa[3]
                R18EC103F= cellsa[4]
                R18EC104F= cellsa[5]
                R18EC105C= cellsa[6]
                R18EC106P= cellsa[7]
                R18EC107P= cellsa[8]
                R18EC108P= cellsa[9]
                R18EC109P= cellsa[10]
                R18EC110P= cellsa[11]
                RUBRICS= cellsa[12]
                CREDITS= cellsa[14]
                TOTALmARKS= cellsa[13]
                TOTALGRADE= cellsa[15]
                SGPA= cellsa[16]
                CGPA= cellsa[17]
                RESULT = cellsa[18]
                #print(f'{i+1},{PIN}\n,{NAME}\n,{R18EC101F}')#R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P,RUBRICS,CREDITS,TOTALGRADE,SGPA,CGPA,RESULT,TOTALmARKS)
                exfill(Marks_sheet,i+1,PIN,NAME,R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P,RUBRICS,CREDITS,TOTALGRADE,SGPA,CGPA,RESULT,TOTALmARKS)
                exfill_old(Marks_sheet_old,i,PIN, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P)
                exfill_old2(Marks_sheet_old2,i,PIN, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P)
                
                i=i+1
    os.makedirs(f'OUTPUTS', exist_ok=True)
    wb.save(f'OUTPUTS/{fname}_SplitOnly.xlsx')

    wb_old.save(f'OUTPUTS/{fname}_MidsAsExt_ForNBA_Criteria3.xlsx')
    wb_old2.save(f'OUTPUTS/{fname}_MidsAsInt_ForNBA_Criteria3.xlsx')

def pd_splitc(a):
    b = a.replace('\n','')
    b = b.replace(' ','')
    b = b.split('(')
    t = b[2]
    t = t.split(')')[0]
    a = a.split('(')
    a = a[1]
    a = a.split(')')
    a = a[0]
    a = a.split("+")
    return a[0],a[1],a[2],a[3],t,b[0]

def splith(a):
    b = a.split('(')[0]
    b = b.replace('\n','')
    b = b.replace(' ','')
    
    return b
    
def exfill(Marks_sheet,i,PIN,NAME, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P,RUBRICS,CREDITS,TOTALGRADE,SGPA,CGPA,RESULT,TOTALmARKS):
    row_pad =1
    Marks_sheet.cell(row=i+row_pad, column=1).value = int(i-1)
    Marks_sheet.cell(row=i+row_pad, column=2).value = str(PIN)
    Marks_sheet.cell(row=i+row_pad, column=3).value = str(NAME)
    print(i-1, PIN,end=' ')
    col =4
    for sub in (R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
        #print(sub)
        mid1,mid2,intr,ext,total,grade = pd_splitc(sub)#splitc(sub)
        #print(mid1,mid2,intr,ext,total,grade)
        try:
            mid1 =float(mid1)
        except:
            mid1 =0

        try:
            mid2 =float(mid2)
        except:
            mid2 =0
        try:
            intr =float(intr)
        except:
            intr =0
        try:
            ext =float(ext)
        except:
            ext =0
        total = mid1+mid2+intr+ext
        status = "PASS" if grade !='E' else "Fail"

        
        for stat in [mid1,mid2,intr,ext,total,grade,status]:
            
            Marks_sheet.cell(row=i+row_pad, column=col).value = stat
            print(stat, end = ' ')
            col=col+1
            
        
    for stat in [RUBRICS,CREDITS,TOTALmARKS,TOTALGRADE,SGPA,CGPA,RESULT]:
                try:
                    stat =float(stat)
                except:
                    stat =stat
        
                Marks_sheet.cell(row=i+row_pad, column=col).value = stat
                col=col+1
                
    print('')

def exfill_old(Marks_sheet_old,i,PIN, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
    
    Marks_sheet_old.cell(row=i+5, column=1).value = int(i)
    Marks_sheet_old.cell(row=i+5, column=2).value = str(PIN)
    #print(i, PIN,end=' ')
    pad =2
    j=0
    for x in (R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
        mid1,mid2,intr,ext,total,grade = pd_splitc(x)#splitc(x)
        try:
            mid1 =float(mid1)
        except:
            mid1 =0

        try:
            mid2 =float(mid2)
        except:
            mid2 =0
        try:
            intr =float(intr)
        except:
            intr =0
        try:
            ext =float(ext)
        except:
            ext =0
        EXTERN_EXAM = float(mid1)+float(mid2)+float(ext)
        INTERN_EXAM = float(intr)
        
        
        Marks_sheet_old.cell(row=i+5, column=pad+1+7*j).value = mid1
        Marks_sheet_old.cell(row=i+5, column=pad+2+7*j).value = mid2
        Marks_sheet_old.cell(row=i+5, column=pad+3+7*j).value = ext
        Marks_sheet_old.cell(row=i+5, column=pad+4+7*j).value = EXTERN_EXAM#get_column_letter(1)
        Marks_sheet_old.cell(row=i+5, column=pad+5+7*j).value = f'=IF({get_column_letter(pad+4+7*j)}{i+5}>={get_column_letter(pad+4+7*j+1)}$5,"Y","N")'
        Marks_sheet_old.cell(row=5, column=pad+5+7*j).value = f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+5+7*j-1)}6:{get_column_letter(pad+5+7*j-1)}{i+5}),0)'
        #print(f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+4+7*j-1)}5:{get_column_letter(pad+6+7*j-1)}{i+5},0))')
        #Marks_sheet_old.cell(row=i+4, column=pad+5+7*j).value =  EX
        Marks_sheet_old.cell(row=i+5, column=pad+6+7*j).value = INTERN_EXAM
        Marks_sheet_old.cell(row=i+5, column=pad+7+7*j).value = f'=IF({get_column_letter(pad+6+7*j)}{i+5}>={get_column_letter(pad+6+7*j+1)}$5,"Y","N")'
        Marks_sheet_old.cell(row=5, column=pad+7+7*j).value = f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+7+7*j-1)}6:{get_column_letter(pad+7+7*j-1)}{i+5}),0)'
        #Marks_sheet_old.cell(row=i+4, column=pad+7+7*j).value = INTX
        j=j+1
        #print(f'{mid1}, {mid2}, {intr}, {ext}', end=' ')

    #print('')

def exfill_old2(Marks_sheet_old,i,PIN, R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
    
    Marks_sheet_old.cell(row=i+5, column=1).value = int(i)
    Marks_sheet_old.cell(row=i+5, column=2).value = str(PIN)
    #print(i, PIN,end=' ')
    pad =2
    j=0
    for x in (R18EC101F,R18EC102F,R18EC103F,R18EC104F,R18EC105C,R18EC106P,R18EC107P,R18EC108P,R18EC109P,R18EC110P):
        mid1,mid2,intr,ext,total,grade = pd_splitc(x)#splitc(x)
        try:
            mid1 =float(mid1)
        except:
            mid1 =0

        try:
            mid2 =float(mid2)
        except:
            mid2 =0
        try:
            intr =float(intr)
        except:
            intr =0
        try:
            ext =float(ext)
        except:
            ext =0
        #EXTERN_EXAM = float(mid1)+float(mid2)+float(ext)
        #INTERN_EXAM = float(intr)

        INTERN_EXAM = float(mid1)+float(mid2)+float(intr)
        EXTERN_EXAM = float(ext)
        
        
        Marks_sheet_old.cell(row=i+5, column=pad+1+7*j).value = mid1
        Marks_sheet_old.cell(row=i+5, column=pad+2+7*j).value = mid2
        Marks_sheet_old.cell(row=i+5, column=pad+3+7*j).value = intr
        Marks_sheet_old.cell(row=i+5, column=pad+4+7*j).value = INTERN_EXAM#get_column_letter(1)
        Marks_sheet_old.cell(row=i+5, column=pad+5+7*j).value = f'=IF({get_column_letter(pad+4+7*j)}{i+5}>={get_column_letter(pad+4+7*j+1)}$5,"Y","N")'
        Marks_sheet_old.cell(row=5, column=pad+5+7*j).value = f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+5+7*j-1)}6:{get_column_letter(pad+5+7*j-1)}{i+5}),0)'
        #print(f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+4+7*j-1)}5:{get_column_letter(pad+6+7*j-1)}{i+5},0))')
        #Marks_sheet_old.cell(row=i+4, column=pad+5+7*j).value =  EX
        Marks_sheet_old.cell(row=i+5, column=pad+6+7*j).value = EXTERN_EXAM
        Marks_sheet_old.cell(row=i+5, column=pad+7+7*j).value = f'=IF({get_column_letter(pad+6+7*j)}{i+5}>={get_column_letter(pad+6+7*j+1)}$5,"Y","N")'
        Marks_sheet_old.cell(row=5, column=pad+7+7*j).value = f'=ROUNDDOWN(AVERAGE({get_column_letter(pad+7+7*j-1)}6:{get_column_letter(pad+7+7*j-1)}{i+5}),0)'
        #Marks_sheet_old.cell(row=i+4, column=pad+7+7*j).value = INTX
        j=j+1
        #print(f'{mid1}, {mid2}, {intr}, {ext}', end=' ')

    #print('')

def readxl(file_name,outdir,andir):
        #print(f"Reading {file_name}.......",end='')
        df = pd.read_excel(file_name, header=[1], sheet_name="Sheet")
        
        cols  = list(df.columns.values)
        #subList = [x for x in cols if "_" in x and "-" in x  ]
        subList = [x for x in cols if "_" in x ]
        subList = [x.split("_")[0] for x in  subList]
        subList = [*set(subList)]
        subList.sort()
        fnameForAnyal = file_name.split('_SplitOnly.xlsx')[0]
        resAnal_workbook = openpyxl.load_workbook(f'{fnameForAnyal}_ResultAnalysis.xlsx')
        resAnal_sheet = resAnal_workbook.active
        res_row_pad = 7
        res_col_pad = 1
        for stat in ["Sl.No","Course Code","Total Students","No of Passed Students in the Course","No of Failed Students in the Course", "Pass Percentage of the Course", "Name_of_Faculty","Sign_of_Faculty"]:
                resAnal_sheet.cell(row=res_row_pad, column=res_col_pad).value = stat
                res_col_pad=res_col_pad+1 
        for sub in subList:
            writer = pd.ExcelWriter(f'{outdir}/{sub}.xlsx')
            col = [ x for x in cols if sub in x] 
            #print(col)
            df_over = df[["PIN","NAME"]+col]
            df_over.insert(loc=0, column="Sl.No.", value=df_over.reset_index().index+1)
            df_over.to_excel(writer,sheet_name=f'{sub}_overview',index=False)
            res_row_pad = res_row_pad +1
            res_col_pad = 1
            for c in col:                
                if "_status" in c:
                    rslt_df = df
                    pas = rslt_df[f'{c}'].value_counts()['PASS']
                    total = rslt_df[f'{c}'].count()
                    fail =  total - pas
                    data_dict = {f"{sub}":["Total Students:","No of Passed Students:","No of Failed Students:", "Pass Percentage:"],
                            f"Result Analysis":[f"{total}",f"{pas}",f"{fail}", f"{round((pas/total)*100,2)}"]}                    
                    data  = pd.DataFrame(data_dict)
                    data.to_excel(writer,sheet_name=f'{sub}_Result Analysis',index=False)
                    subcodeno_forAnal = sub[-1] if sub[-1]!="0" else "10"
                    resAnal_sheet.cell(row=res_row_pad, column=res_col_pad).value = int(subcodeno_forAnal)
                    resAnal_sheet.cell(row=res_row_pad, column=res_col_pad+1).value = f'{sub}'
                    for stat in data_dict[ f"Result Analysis"]:
                        resAnal_sheet.cell(row=res_row_pad, column=res_col_pad+2).value = stat
                        res_col_pad=res_col_pad+1 
            for c in col:
                if "_mid1" in c:                    
                    rslt_df = df.loc[(df[c] < (0.01* SLOW_LEARNERS_PERCENT * MID1_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'MID1_DULL',index=False)

                    rslt_df = df.loc[(df[c] >= (0.01* FAST_LEARNERS_PERCENT * MID1_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'MID1_TOP',index=False)

                if "_mid2" in c:
                    
                    rslt_df = df.loc[(df[c] < (0.01* SLOW_LEARNERS_PERCENT * MID2_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'MID2_DULL',index=False)

                    rslt_df = df.loc[(df[c] >= (0.01* FAST_LEARNERS_PERCENT * MID2_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'MID2_TOP',index=False)

                if "_intr" in c:
                    
                    rslt_df = df.loc[(df[c] < (0.01* SLOW_LEARNERS_PERCENT * INTERNAL_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'intr_DULL',index=False)

                    rslt_df = df.loc[(df[c] >= (0.01* FAST_LEARNERS_PERCENT * INTERNAL_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'intr_TOP',index=False)
                if "_ext" in c:
                    
                    rslt_df = df.loc[(df[c] < (0.01* SLOW_LEARNERS_PERCENT * EXTERNAL_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'ext_DULL',index=False)

                    rslt_df = df.loc[(df[c] >= (0.01* FAST_LEARNERS_PERCENT * EXTERNAL_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'ext_TOP',index=False)
                if "_total" in c:
                    
                    rslt_df = df.loc[(df[c] < (0.01* SLOW_LEARNERS_PERCENT * TOTAL_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index+1)
                    rslt_df.to_excel(writer,sheet_name=f'total_DULL',index=False)

                    rslt_df = df.loc[(df[c] >= (0.01* FAST_LEARNERS_PERCENT * TOTAL_MAX_MARKS ))]
                    rslt_df = rslt_df[["PIN","NAME",f'{c}']]
                    rslt_df.insert(loc=0, column="Sl.No.", value=rslt_df.reset_index().index + 1)
                    rslt_df.to_excel(writer,sheet_name=f'total_TOP',index=False)


            writer.close()
        #resAnal_sheet.merge_cells(start_row=res_row_pad+2, start_column=6, end_row=res_row_pad+2, end_column=8)
        resAnal_sheet.cell(row=res_row_pad+2, column=6).value = f'Signature of HOD'
        resAnal_sheet.insert_rows(1)
        resAnal_sheet.insert_rows(1)
        resAnal_sheet.insert_rows(1)
        resAnal_sheet.insert_rows(1)
        resAnal_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        resAnal_sheet.cell(row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
        resAnal_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        resAnal_sheet.cell(row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
        resAnal_sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
        resAnal_sheet.cell(row=4, column=1).value = "Exam Month and Year:"
        resAnal_sheet.cell(row=5, column=1).value = "Sl.No"
        resAnal_sheet.insert_rows(10)
        resAnal_sheet.merge_cells(start_row=11, start_column=1, end_row=11, end_column=8)
        resAnal_sheet.cell(row=11, column=1).value = "CourseWise Analysis"
        resAnal_workbook.save(f'{fnameForAnyal}_ResultAnalysis.xlsx')
    
def conect(url):
    timeout = 5
    op = None
    while True:
        try:
            r = requests.get(url,verify = False, timeout=timeout)
            op = r.status_code
            if op == 200:
                pass
                # print("Done.")
                break
            else:
                # print("Status Code is not 200")
                # print("status Code", op)
                pass
        except:
            # print("Not Connected. No internet.")
            # print("trying again...")
            sleep(5)

    return r

def conect_with_print(url):
    timeout = 5
    op = None
    #print(url)
    while True:
        try:
            r = requests.get(url, verify = False, timeout=timeout)
            op = r.status_code
            #print(r)
            if op == 200:
                pass
                print("Done.")
                break
            else:
                print("Status Code is not 200")
                print("status Code", op)
                pass
        except Exception as e:
            #print(e)
            #print("status Code", op)
            print('''Unable to Connect to server.
Please check your internet connection.
if internet is okay.
Contact keerthichand.c@gmail.com''')
            print("trying again...")
            sleep(5)
    return r

def string_toFloat(x):
    try:
        return float(x)

    except:
        return x

def non_zero(x):

    try:
        return float(x)

    except:
        return 0

def dip_res(PIN):
    url = f'{c18}{PIN}'
    if __name__ == "__main__":
        r = conect_with_print(url)
    else:
        r = conect(url)
    data = r.json()
    data = loads(data)
    # print(data)
    details = {}
    if len(data) == 4:
        # print(data['Table'])
        try:
            details["StudentName"] = data['Table'][0]['StudentName']
            details["Pin"] = data['Table'][0]['Pin']
            details["BranchCode"] = data['Table'][0]['BranchCode']
            details["Scheme"] = data['Table'][0]['Scheme']
        except:
            details["StudentName"] = 'Details Not Found'
            details["Pin"] = 'Details Not Found'
            details["BranchCode"] = 'Details Not Found'
            details["Scheme"] = 'Details Not Found'

        # print(data['Table1'])
        try:
            details["CGPA"] = data['Table1'][0]['CGPA']
            details["CreditsGained"] = data['Table1'][0]['CreditsGained']
        except:
            details["CGPA"] = 'Details Not Found'
            details["CreditsGained"] = 'Details Not Found'
        cred = string_toFloat(details["CreditsGained"])
        cred = cred + len(data['Table3'])*2.5
        if type(cred)== float:
            cred + len(data['Table3'])*2.5
            if cred >= 130:
                details["Diploma_Status"] = "Completed_Diploma"
            else:
                details["Diploma_Status"] = "Not_Completed_Diploma"
        else:
            details["Diploma_Status"] = "Details_not_found"        
            
            

        #details["SGPA_TOTAL"] = data['Table3']
        sem_data = []
        if len(data['Table3']) != 0:
            for sem in data['Table3']:
                sem_pass = []
                sem_stat = []
                # print(f'geting{sem["Semester"]}')
                for exam in data['Table2']:
                    # print(exam,sem)
                    # print(exam['Semester'])
                    if exam['Semester'] == sem['Semester']:
                        # print(exam['Semester'],exam['WholeOrSupply'],exam["ExamStatus"])
                        sem_pass.append(exam['WholeOrSupply'])
                        sem_stat.append(exam['ExamStatus'])
                # print(sem_pass)
                sem_pass = set(sem_pass)
                sem_stat = set(sem_stat)
                passed_stat = "Not Passed"
                passed_data = "SUPPLY"
                # print(sem_stat)
                if len(sem_stat) == 1:
                    if "P" in sem_stat:
                        passed_stat = "passed"
                if len(sem_pass) == 1:
                    if "W" in sem_pass:
                        passed_data = "REGULAR"

                sem["PASSED_TYPE"] = passed_data
                sem["PASSED_STATUS"] = passed_stat
                sem_data.append(sem)

        details['SEM_DETAILS'] = sem_data

    return details, data

def col_adju(ws: openpyxl.worksheet.worksheet.Worksheet):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                #dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value*0.65

def columns_best_fit(ws: openpyxl.worksheet.worksheet.Worksheet):
        """
        Make all columns best fit
        """
        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].bestFit = True

def getList(dict):
    return dict.keys()

def reg_sup(sem_pass):
    return "REGULAR" if "W" in sem_pass else "SUPPLY"

def reg_stat(sem_pass):

    if "FNP" in sem_pass:
        return "Fee Not Paid"
    elif "F" in sem_pass:
        return "Failed"

    elif "P" in sem_pass:
        return "PASSED"
    else:
        return "Status Unkwon"

def insert(sheet_work,YERS,colmnr,rowmax,ROW_PAD):
    sheet_work.insert_cols(colmnr)
    sheet_work.cell(row= ROW_PAD, column=colmnr).value = f'{YERS}'
    # print(rowmax)
    for rowg  in range(ROW_PAD+1,rowmax+1):
        first_cell = sheet_work.cell(row=rowg, column=colmnr-1).coordinate
        secon_cell = sheet_work.cell(row=rowg, column=colmnr-2).coordinate
        third_cell = sheet_work.cell(row=rowg, column=colmnr-5).coordinate
        fourth_cell = sheet_work.cell(row=rowg, column=colmnr-6).coordinate
        #print("Hello")
        # sheet_work.cell(row= rowg, column=colmnr).value = f'=IF({first_cell}="passed",IF({secon_cell}="REGULAR",IF({third_cell}="passed",IF({fourth_cell}="REGULAR","First_attempt","NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt")'
        #print(f'=IF({first_cell}="passed",IF({secon_cell}="REGULAR",IF({third_cell}="passed",IF({fourth_cell}="REGULAR","First_attempt","NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt")')
        sheet_work.cell(row= rowg, column=colmnr).value =f'=IF({first_cell}="passed",IF({third_cell}="passed",IF({secon_cell}="REGULAR",IF({fourth_cell}="REGULAR","First_attempt","NOT_First_attempt"),"NOT_First_attempt"),"Not_Passed"),"Not_Passed")'
        
    firstcell_P = sheet_work.cell(row=ROW_PAD+1, column=colmnr).coordinate
    passedcell_P = sheet_work.cell(row=rowmax, column=colmnr).coordinate
    
    sheet_work.cell(row=rowmax+1, column=colmnr).value = f'No Students appeared'
    sheet_work.cell(row=rowmax+2, column=colmnr).value = f'=COUNTIF({firstcell_P}:{passedcell_P},"<>"&"")'
    sheet_work.cell(row=rowmax+3, column=colmnr).value = f'No Students passed without Backlogs(First Attempt)'
    sheet_work.cell(row=rowmax+4, column=colmnr).value = f'=COUNTIF({firstcell_P}:{passedcell_P},"First_attempt")'
    sheet_work.cell(row=rowmax+5, column=colmnr).value = f'No Students passed (with or without Backlogs)'
    sheet_work.cell(row=rowmax+6, column=colmnr).value = f'=COUNTIF({firstcell_P}:{passedcell_P},"<>"&"")-COUNTIF({firstcell_P}:{passedcell_P},"Not_Passed")'

##    sheet_work.cell(row=rowmax+1, column=colmnr).alignment = Alignment(wrap_text=True)
##    sheet_work.cell(row=rowmax+2, column=colmnr).alignment = Alignment(wrap_text=True)
##    sheet_work.cell(row=rowmax+3, column=colmnr).alignment = Alignment(wrap_text=True)
##    sheet_work.cell(row=rowmax+4, column=colmnr).alignment = Alignment(wrap_text=True)
##    sheet_work.cell(row=rowmax+5, column=colmnr).alignment = Alignment(wrap_text=True)
##    sheet_work.cell(row=rowmax+6, column=colmnr).alignment = Alignment(wrap_text=True)

    firstcell_P = sheet_work.cell(row=rowmax+2, column=colmnr).coordinate
    passedcell_P = sheet_work.cell(row=rowmax+4, column=colmnr).coordinate
    passedcell_wP = sheet_work.cell(row=rowmax+6, column=colmnr).coordinate
    sheet_work.cell(row=rowmax+7, column=colmnr).value = f'No Students passed without Backlogs(First Attempt) Percentage'
    sheet_work.cell(row=rowmax+8, column=colmnr).value = f"=ROUND({passedcell_P}/{firstcell_P}*100,1)"
    sheet_work.cell(row=rowmax+9, column=colmnr).value = f'No Students passed (with or without Backlogs) Percentage'
    sheet_work.cell(row=rowmax+10, column=colmnr).value = f"=ROUND({passedcell_wP}/{firstcell_P}*100,1)"

def last_insert(sheet_work,colmnr,rowmax,ROW_PAD):
    # colmnr = sheet_work.max_coulmn
    #sheet_work.insert_cols(colmnr)
    sheet_work.cell(row= ROW_PAD, column=colmnr).value = f'DIPLOMA_NO_BACKLOG'
    #print(rowmax)
    for rowg  in range(ROW_PAD+1,rowmax+1):
        dipstatus_cell =  sheet_work.cell(row=rowg, column=len(TOP_ROW_gen)).coordinate
        first_cell = sheet_work.cell(row=rowg, column=colmnr-1).coordinate
        secon_cell = sheet_work.cell(row=rowg, column=colmnr-10).coordinate
        third_cell = sheet_work.cell(row=rowg, column=colmnr-19).coordinate
        # fourth_cell = sheet_work.cell(row=rowg, column=colmnr-6).coordinate
        #sheet_work.cell(row= rowg, column=colmnr).value = f'=IF({first_cell}="First_attempt",IF({secon_cell}="First_attempt",IF({third_cell}="First_attempt","PASSED_WITH_OUT_BACK","BACKLOG"),"BACKLOG"),"BACKLOG")'
        sheet_work.cell(row= rowg, column=colmnr).value = f'=(IF({dipstatus_cell}="Completed_Diploma",IF({first_cell}="First_attempt",IF({secon_cell}="First_attempt",IF({third_cell}="First_attempt","WITHOUT_BACKLOG_PASSED_Diploma","WITH_BACKLOG_PASSED_Diploma"),"WITH_BACKLOG_PASSED_Diploma"),"WITH_BACKLOG_PASSED_Diploma"),"NOT_PASSED_Diploma"))'
        #print(rowg)

def cal_per(sheet,SEM,row_max,ROW_PAD):
    
    for count in SEM_CONT:
        colr = SEM-count
        sheet.cell(row=row_max+1, column=colr-1).value = f"No Students appeared"
        firstcell = sheet.cell(row=ROW_PAD+1, column=colr-1).coordinate
        passedcell = sheet.cell(row=row_max, column=colr-1).coordinate
        sheet.cell(row=row_max+1, column=colr).value = f'=COUNTIF({firstcell}:{passedcell},"<>"&"")'
        # total = sheet.cell(row=row_max+2, column=colr).coordinate

        firstcell_P = sheet.cell(row=ROW_PAD+1, column=colr).coordinate
        passedcell_P = sheet.cell(row=row_max, column=colr).coordinate

        sheet.cell(row=row_max+2, column=colr).value = f'=COUNTIFS({firstcell}:{passedcell},"REGULAR",{firstcell_P}:{passedcell_P},"passed")'
        passed = sheet.cell(row=row_max+2, column=colr).coordinate

        sheet.cell(row=row_max+2, column=colr -1).value = "No of students Passed in Regular"

        sheet.cell(row=row_max+3, column=colr -1).value = "Regular Pass Percentage"
        total = sheet.cell(row=row_max+1, column=colr).coordinate
        sheet.cell(row=row_max+3, column=colr).value = f"=ROUND({passed}/{total}*100,1)"

def fill_excel_total(exam, sheet, col, row, ROW_PAD):

    # print(col)
    Subject_Code = exam["Subject_Code"]
    Subject = exam

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['Mid1Marks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_Mid1Marks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['Mid2Marks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_Mid2Marks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['InternalMarks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_InternalMarks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['EndExamMarks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_EndExamMarks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['SubjectTotal']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_SubjectTotal"
    col = col+1

    sheet.cell(row=row+1, column=col).value = f"{Subject['ExamMonthYear']}"
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_ExamMonthYear"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = f"{reg_sup(Subject['WholeOrSupply'])}"
    sheet.cell(
        row=ROW_PAD, column=col).value = f"{Subject_Code}_Writing_Regular_or_Suplly"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = f"{reg_stat(Subject['ExamStatus'])}"
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_ExamStatus"

    sheet.cell(row=row+1+1, column=col-1).value = f" No of Students appeared"
    firstcell = sheet.cell(row=ROW_PAD+1, column=col-1).coordinate
    passedcell = sheet.cell(row=row+1, column=col-1).coordinate
    sheet.cell(
        row=row+1+1, column=col).value = f'=COUNTIF({firstcell}:{passedcell},"<>"&"")'
    total = sheet.cell(row=row+1+1, column=col).coordinate

    firstcell_P = sheet.cell(row=ROW_PAD+1, column=col).coordinate
    passedcell_P = sheet.cell(row=row+1, column=col).coordinate

    sheet.cell(
        row=row+1+2, column=col).value = f'=COUNTIFS({firstcell}:{passedcell},"REGULAR",{firstcell_P}:{passedcell_P},"PASSED")'
    passed = sheet.cell(row=row+1+2, column=col).coordinate

    sheet.cell(row=row+1+2, column=col -
               1).value = " No of students Passed in Regular"

    sheet.cell(row=row+1+3, column=col-1).value = "Regular Pass Percentage"
    sheet.cell(
        row=row+1+3, column=col).value = f"=ROUND({passed}/{total}*100,1)"

    col = col+1

    return col

def fill_excel_overview(details, sheet, row, ROW_PAD):
    col = 1
    for key in getList(details):
        if key != 'SEM_DETAILS':
            sheet.cell(
                row=row+1, column=col).value = string_toFloat(f'{details[key]}')
            col = col+1
        if key == 'SEM_DETAILS':
            for exam in details['SEM_DETAILS']:
                sheet.cell(
                    row=row+1, column=col).value = string_toFloat(f"{exam['Credits']}")
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_credits"
                col = col+1

                sheet.cell(
                    row=row+1, column=col).value = string_toFloat(f"{exam['SGPA']}")
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_SGPA"
                col = col+1

                sheet.cell(
                    row=row+1, column=col).value = f"{exam['PASSED_TYPE']}"
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_PASSED_TYPE"
                
                col = col+1

                sheet.cell(
                    row=row+1, column=col).value = f"{exam['PASSED_STATUS']}"
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_PASSED_STATUS"
                
                
            

                

                col = col+1
                    
def getresult_Excel(pins):
    
    wb_obj = openpyxl.Workbook()
    sheet_over = wb_obj.create_sheet(index=0, title="Overview")
    sem_sheets = {}
    for i in range(6):
        sem_sheets[f"{i+1}SEM"] = wb_obj.create_sheet(
            index=i+1, title=f"{i+1}SEM")
        col = 0
        ROW_PAD = 3
        sem_sheets[f"{i+1}SEM"].cell(
            row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
        sem_sheets[f"{i+1}SEM"].cell(
            row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
        for trow in TOP_ROW_tot:
            sem_sheets[f"{i+1}SEM"].cell(row=ROW_PAD,
                                         column=col+1).value = f'{trow}'
            col = col+1

    col = 0
    ROW_PAD = 3
    sheet_over.cell(
        row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    sheet_over.cell(
        row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"

    for trow in TOP_ROW_gen:
        sheet_over.cell(row=ROW_PAD, column=col+1).value = f'{trow}'
        col = col+1

    for i in range(len(pins)):
        if __name__ == "__main__":
            print(f"geting result data for PIN: {pins[i]}....", end='')
        details, data = dip_res(pins[i])
        # print(details)
        fill_excel_overview(details, sheet_over, i+ROW_PAD, ROW_PAD)
        col = 0
        try:
            for key in getList(sem_sheets):
                col = 1
                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["Scheme"]}'
                col = col+1
                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["StudentName"]}'
                col = col+1
                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["Pin"]}'
                col = col+1

                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["BranchCode"]}'
                col = col+1
                try:
                    for exam in data["Table2"]:
                        if key == exam['Semester']:
                           # print(col)

                            col = fill_excel_total(exam, sem_sheets[key], col, i+ROW_PAD, ROW_PAD)
                            #print("excuting")

                except:
                    sem_sheets[key].cell(
                        row=i+ROW_PAD+1, column=col).value = f'Details  not found'
                    col = col+1

        except:
            sem_sheets[key].cell(
                row=i+ROW_PAD+1, column=3).value = f'{pins[i]}'
            col = col+1

    
    row_max = sheet_over.max_row
    
    
    for YEARS in SEM_INSERT.keys():
        insert(sheet_over,YEARS,SEM_INSERT[YEARS],row_max,ROW_PAD)
    
    col_max = sheet_over.max_column + 1  
    last_insert(sheet_over,col_max,row_max,ROW_PAD)

    firstcell_o = sheet_over.cell(row=ROW_PAD+1, column=len(TOP_ROW_gen)).coordinate
    passedcell_o = sheet_over.cell(row_max, column=len(TOP_ROW_gen)).coordinate
    
    sheet_over.cell(row=row_max+1, column=len(TOP_ROW_gen)-1).value = f"No Diploma Completed Students:"
    sheet_over.cell(row=row_max+1, column=len(TOP_ROW_gen)).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"Completed_Diploma")'

    firstcell_o = sheet_over.cell(row=ROW_PAD+1, column=col_max).coordinate
    passedcell_o = sheet_over.cell(row_max, column=col_max).coordinate

    sheet_over.cell(row=row_max+1, column=col_max).value = f"No Students:"
    sheet_over.cell(row=row_max+1, column=col_max+1).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"<>"&"")'
    
    sheet_over.cell(row=row_max+2, column=col_max).value = f"No Diploma Completed without Backlogs Students:"
##    sheet_over.cell(row=row_max+2, column=col_max).alignment = Alignment(wrap_text=True)
    sheet_over.cell(row=row_max+2, column=col_max+1).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"WITHOUT_BACKLOG_PASSED_Diploma")'

    sheet_over.cell(row=row_max+3, column=col_max).value = f"No Diploma Completed with Backlogs Students:"
##    sheet_over.cell(row=row_max+3, column=col_max).alignment = Alignment(wrap_text=True)
    sheet_over.cell(row=row_max+3, column=col_max+1).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"WITH_BACKLOG_PASSED_Diploma")'

    sheet_over.cell(row=row_max+4, column=col_max).value = f"No Diploma Completed Students (with or without Backlogs):"
##    sheet_over.cell(row=row_max+4, column=col_max).alignment = Alignment(wrap_text=True)
    sheet_over.cell(row=row_max+4, column=col_max+1).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"WITH_BACKLOG_PASSED_Diploma")+COUNTIF({firstcell_o}:{passedcell_o},"WITHOUT_BACKLOG_PASSED_Diploma")'


    ns_count = sheet_over.cell(row=row_max+1, column=col_max+1).coordinate
    wob_count = sheet_over.cell(row=row_max+2, column=col_max+1).coordinate
    wb_count = sheet_over.cell(row=row_max+3, column=col_max+1).coordinate
    dc_count = sheet_over.cell(row=row_max+4, column=col_max+1).coordinate
    
    sheet_over.cell(row=row_max+5, column=col_max).value = f"Without BackLog Percentage"
    sheet_over.cell(row=row_max+5, column=col_max+1).value =f'=ROUND({wob_count}/{ns_count}*100,1)'

    sheet_over.cell(row=row_max+6, column=col_max).value = f"WithBackLog Percentage"
    sheet_over.cell(row=row_max+6, column=col_max+1).value =f'=ROUND({wb_count}/{ns_count}*100,1)'

    sheet_over.cell(row=row_max+7, column=col_max).value = f"Diploma Completed Percentage"
    sheet_over.cell(row=row_max+7, column=col_max+1).value =f'=ROUND({dc_count}/{ns_count}*100,1)'
    
    
    sheet_disc = wb_obj.create_sheet(index=0, title="Disclaimer")
    for YEARS in SEM_INSERT.keys():
        cal_per(sheet_over,SEM_INSERT[YEARS],row_max,ROW_PAD)
    
    sheet_disc.cell(row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    sheet_disc.cell(row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
    
    
    sheet_disc.merge_cells('A3:AB5') 
    sheet_disc.cell(row=3, column=1).value = '''
    
    


disclaimer:
    *These are not actual Results
    *Use this at your own discretion
    *for actual Results please visit TS-SBTET Website
    *OUTPUTS from this program are not related to TS-SBTET
    *This program is developed only to help ease work but not to misguide or harm in 
        any possible ways and with no ill intent
        '''
        

    
    return wb_obj

def get_cons(pins,fname):
    wb_obj = getresult_Excel(pins)
    for ws in wb_obj.sheetnames:
        col_adju(wb_obj[ws])
        columns_best_fit(wb_obj[ws])
        i = wb_obj[ws].max_column
        wb_obj[ws].column_dimensions[get_column_letter(i)].width = 5
        try:
            wb_obj[ws].column_dimensions[get_column_letter(i-1)].width = 54
        except:
            pass
        wb_obj[ws].column_dimensions['B'].width = 15
        wb_obj[ws].column_dimensions['E'].width = 5.71
        wb_obj[ws].column_dimensions['A'].width = 28
        wb_obj[ws].column_dimensions['P'].width = 60
        wb_obj[ws].column_dimensions['Y'].width = 60
        wb_obj[ws].column_dimensions['AH'].width = 60
        
    os.makedirs(f"{file_path}\\OUTPUTs", exist_ok=True)
    wb_obj.save(f"{file_path}\\OUTPUTs\\{fname}_consolidated_AsOn_{TODAY}.xlsx")
   
if __name__=='__main__':
    count = 0
    coderintro()
    instructions_info()
    dir_list = os.listdir(f'{thisdir}/INPUT_HTML')
    os.makedirs(f'OUTPUTS', exist_ok=True)
    for file in dir_list:
        if (file.endswith(".html") or file.endswith(".htm")) :
            #print(file)
            print(f"Converting {file}")
            #readht(file)
            PD_readht(file)
            print(f"Conversion copleted.")
            count +=1
    
    if count >0:
        print("\n\n\n")
        print("Now subject wise Analysis......")
        dir_list = os.listdir(f'{thisdir}/OUTPUTS')
        for file in dir_list:
            if (file.endswith("_SplitOnly.xlsx")): #or file.endswith(".xlx")) :
                #print(f'{thisdir}/OUTPUTS/{file}')
                print(f"Analysing {file}.........",end = " ")
                os.makedirs(f'OUTPUTS/SUBJECT_WISE_ANALYSIS', exist_ok=True)
                readxl(f'{thisdir}/OUTPUTS/{file}',f'{thisdir}/OUTPUTS/SUBJECT_WISE_ANALYSIS',f'{thisdir}/OUTPUTS')
                print(f"Analysis copleted.")

        print("All files are saved in Outputs folder")
        input("Press Enter to Exit")
    else:
        print("No html files found.......")
        print("please read following instructions")
        instructions_info()
        print("Please Start the program again with proper inputs...")
        input("Press Enter to Exit")
                
                
    
