import os
from time import sleep
import openpyxl
import urllib3
urllib3.disable_warnings()
import requests


from json import loads

c16s = "https://www.sbtet.telangana.gov.in/API/api/Results/GetC16SConsolidatedResult?Pin="
c18 = "https://www.sbtet.telangana.gov.in/API/api/Results/GetConsolidatedResults?Pin="
c21 = "https://www.sbtet.telangana.gov.in/API/api/Results/GetConsolidatedResults?Pin="
c16 = "https://www.sbtet.telangana.gov.in/API/api/Results/GetC16ConsolidatedResult?Pin="


TOP_ROW_tot = ['Scheme', 'StudentName', 'Pin', 'BranchCode']
TOP_ROW_SEM = ['Credits', 'SGPA', 'PASSED_TYPE', 'PASSED_STATUS']
TOP_ROW_gen = ['StudentName', 'Pin', 'BranchCode',
               'Scheme', 'CGPA', 'TOTAL_CreditsGained', 'Diploma_Status']

file_path = os.getcwd()

SEM_INSERT = {"1st_Year_Passed":16, 
              "2nd_Year_Passed":25,
              "3rd_Year_Passed":34}

SEM_CONT = [1,5]

def coderintro():

    screen_time =4
    for i in range(0,screen_time):
                os.system('cls')
                print()
                print("********************************************************************************")
                print()
                print("Program for consolidated results and save in excel for C18 and C21 Scheme")
                print("********************************************************************************")
                print()
                print("********************************************************************************")
                print()
                print("developed by Keerthi Chandra C")
                print("Lecturer in ECE , GMRPW Karimnagar")
                print("for constructive criticism and suggestions reach me at keerthichand.c@gmail.com")
                print('''


disclaimer:
    *These are not actual Results
    *Use this at your own discretion
    *for actual Results please visit TS-SBTET Website
    *OUTPUTS from this program are not related to TS-SBTET
    *This program is developed only to help ease work but not to misguide or harm in 
        any possible ways and with no ill intent
    ''')
                print("******************************************************************************")
                if screen_time-i-1>0:
                        print(f'starting in {screen_time-i-2}s .......')
                sleep(1)
    input("press ENTER to start.....")
    os.system('cls')

def instructions_info():
        print('''
************************************************************************************

before staring ...
Paste C18 PINs in INPUT excel fie in Column A in INPUT Folder 
Save and close INPUT excel in INPUT Folder 

************************************************************************************''')
        input("press ENTER to continue......")
        os.system('cls')
        print()

def getpins():
    print(f"getiing PINS from Input sheet...",end='')
    path = f"{file_path}\\INPUT\\INPUT.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    pins = []
    for row in range(1,sheet_obj.max_row+1):
        pin = sheet_obj.cell(row = row, column = 1).value
        #print(pin)
        
        if not(pin == None) and '-' in pin:
            pin = pin.upper()
            pin = pin.replace("\n","")
            pin = pin.replace(" ","")
            pins.append(pin)
            
    #for pin in pins:
        #print(pin)
    wb_obj.close()
    print(f'{len(pins)} found.')
    return pins
    


def check_pin(pins):
    pins=pins[0]
    if " " in pins:
        pins = [pin for pin in pins.split(" ")]
    elif "," in pins:
        pins = [pin for pin in pins.split(",")]
    else:
        pins = [pin for pin in pins.split("\n")]
        
            
    
    
     
    pins = [pin for pin in pins if "-" in pin ]
    pins = [pin.replace(",", "")for pin in pins ]
    pins = [pin.replace("\n", "")for pin in pins ]
    pins = [pin.replace(" ", "")for pin in pins ]
    pins = [pin for pin in pins if len(pin.split("-"))==3]
    pins = [pin for pin in pins if len(pin.split("-")[0])==5]
    pins = [pin for pin in pins if len(pin.split("-")[2])==3]
    pins = [pin for pin in pins if (len(pin)>10 and len(pin)<15) ]
    pins = [pin.upper() for pin in pins]
    return pins
    
def conect(url):
    timeout = 5
    op = None
    while True:
        try:
            r = requests.get(url,verify = False, timeout=timeout)
            op = r.status_code
            if op == 200:
                pass
                # print("Done.")
                break
            else:
                # print("Status Code is not 200")
                # print("status Code", op)
                pass
        except:
            # print("Not Connected. No internet.")
            # print("trying again...")
            sleep(5)

    return r
def conect_with_print(url):
    timeout = 5
    op = None
    #print(url)
    while True:
        try:
            r = requests.get(url, verify = False, timeout=timeout)
            op = r.status_code
            #print(r)
            if op == 200:
                pass
                print("Done.")
                break
            else:
                print("Status Code is not 200")
                print("status Code", op)
                pass
        except Exception as e:
            #print(e)
            #print("status Code", op)
            print('''Unable to Connect to server.
Please check your internet connection.
if internet is okay.
Contact keerthichand.c@gmail.com''')
            print("trying again...")
            sleep(5)
    return r


def string_toFloat(x):
    try:
        return float(x)

    except:
        return x


def non_zero(x):

    try:
        return float(x)

    except:
        return 0


def dip_res(PIN):
    url = f'{c18}{PIN}'
    if __name__ == "__main__":
        r = conect_with_print(url)
    else:
        r = conect(url)
    data = r.json()
    data = loads(data)
    # print(data)
    details = {}
    if len(data) == 4:
        # print(data['Table'])
        try:
            details["StudentName"] = data['Table'][0]['StudentName']
            details["Pin"] = data['Table'][0]['Pin']
            details["BranchCode"] = data['Table'][0]['BranchCode']
            details["Scheme"] = data['Table'][0]['Scheme']
        except:
            details["StudentName"] = 'Details Not Found'
            details["Pin"] = 'Details Not Found'
            details["BranchCode"] = 'Details Not Found'
            details["Scheme"] = 'Details Not Found'

        # print(data['Table1'])
        try:
            details["CGPA"] = data['Table1'][0]['CGPA']
            details["CreditsGained"] = data['Table1'][0]['CreditsGained']
        except:
            details["CGPA"] = 'Details Not Found'
            details["CreditsGained"] = 'Details Not Found'
        cred = string_toFloat(details["CreditsGained"])
        if type(cred)== float:
            if cred >= 130:
                details["Diploma_Status"] = "Completed_Diploma"
            else:
                details["Diploma_Status"] = "Not_Completed_Diploma"
        else:
            details["Diploma_Status"] = "Details_not_found"        
            
            

        #details["SGPA_TOTAL"] = data['Table3']
        sem_data = []
        if len(data['Table3']) != 0:
            for sem in data['Table3']:
                sem_pass = []
                sem_stat = []
                # print(f'geting{sem["Semester"]}')
                for exam in data['Table2']:
                    # print(exam,sem)
                    # print(exam['Semester'])
                    if exam['Semester'] == sem['Semester']:
                        # print(exam['Semester'],exam['WholeOrSupply'],exam["ExamStatus"])
                        sem_pass.append(exam['WholeOrSupply'])
                        sem_stat.append(exam['ExamStatus'])
                # print(sem_pass)
                sem_pass = set(sem_pass)
                sem_stat = set(sem_stat)
                passed_stat = "Not Passed"
                passed_data = "SUPPLY"
                # print(sem_stat)
                if len(sem_stat) == 1:
                    if "P" in sem_stat:
                        passed_stat = "passed"
                if len(sem_pass) == 1:
                    if "W" in sem_pass:
                        passed_data = "REGULAR"

                sem["PASSED_TYPE"] = passed_data
                sem["PASSED_STATUS"] = passed_stat
                sem_data.append(sem)

        details['SEM_DETAILS'] = sem_data

    return details, data


def getList(dict):
    return dict.keys()


def reg_sup(sem_pass):
    return "REGULAR" if "W" in sem_pass else "SUPPLY"


def reg_stat(sem_pass):

    if "FNP" in sem_pass:
        return "Fee Not Paid"
    elif "F" in sem_pass:
        return "Failed"

    elif "P" in sem_pass:
        return "PASSED"
    else:
        return "Status Unkwon"

def insert(sheet_work,YERS,colmnr,rowmax,ROW_PAD):
    sheet_work.insert_cols(colmnr)
    sheet_work.cell(row= ROW_PAD, column=colmnr).value = f'{YERS}'
    # print(rowmax)
    for rowg  in range(ROW_PAD+1,rowmax+1):
        first_cell = sheet_work.cell(row=rowg, column=colmnr-1).coordinate
        secon_cell = sheet_work.cell(row=rowg, column=colmnr-2).coordinate
        third_cell = sheet_work.cell(row=rowg, column=colmnr-5).coordinate
        fourth_cell = sheet_work.cell(row=rowg, column=colmnr-6).coordinate
        #print("Hello")
        # sheet_work.cell(row= rowg, column=colmnr).value = f'=IF({first_cell}="passed",IF({secon_cell}="REGULAR",IF({third_cell}="passed",IF({fourth_cell}="REGULAR","First_attempt","NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt")'
        #print(f'=IF({first_cell}="passed",IF({secon_cell}="REGULAR",IF({third_cell}="passed",IF({fourth_cell}="REGULAR","First_attempt","NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt"),"NOT_First_attempt")')
        sheet_work.cell(row= rowg, column=colmnr).value =f'=IF({first_cell}="passed",IF({third_cell}="passed",IF({secon_cell}="REGULAR",IF({fourth_cell}="REGULAR","First_attempt","NOT_First_attempt"),"NOT_First_attempt"),"Not_Passed"),"Not_Passed")'

def last_insert(sheet_work,colmnr,rowmax,ROW_PAD):
    # colmnr = sheet_work.max_coulmn
    #sheet_work.insert_cols(colmnr)
    sheet_work.cell(row= ROW_PAD, column=colmnr).value = f'DIPLOMA_NO_BACKLOG'
    #print(rowmax)
    for rowg  in range(ROW_PAD+1,rowmax+1):
        dipstatus_cell =  sheet_work.cell(row=rowg, column=len(TOP_ROW_gen)).coordinate
        first_cell = sheet_work.cell(row=rowg, column=colmnr-1).coordinate
        secon_cell = sheet_work.cell(row=rowg, column=colmnr-10).coordinate
        third_cell = sheet_work.cell(row=rowg, column=colmnr-19).coordinate
        # fourth_cell = sheet_work.cell(row=rowg, column=colmnr-6).coordinate
        #sheet_work.cell(row= rowg, column=colmnr).value = f'=IF({first_cell}="First_attempt",IF({secon_cell}="First_attempt",IF({third_cell}="First_attempt","PASSED_WITH_OUT_BACK","BACKLOG"),"BACKLOG"),"BACKLOG")'
        sheet_work.cell(row= rowg, column=colmnr).value = f'=(IF({dipstatus_cell}="Completed_Diploma",IF({first_cell}="First_attempt",IF({secon_cell}="First_attempt",IF({third_cell}="First_attempt","WITHOUT_BACKLOG_PASSED_Diploma","WITH_BACKLOG_PASSED_Diploma"),"WITH_BACKLOG_PASSED_Diploma"),"WITH_BACKLOG_PASSED_Diploma"),"NOT_PASSED_Diploma"))'
        #print(rowg)

def cal_per(sheet,SEM,row_max,ROW_PAD):
    
    for count in SEM_CONT:
        colr = SEM-count
        sheet.cell(row=row_max+1, column=colr-
                           1).value = f"No Students appeared"
        firstcell = sheet.cell(row=ROW_PAD+1, column=colr-1).coordinate
        passedcell = sheet.cell(row=row_max, column=colr-1).coordinate
        sheet.cell(
            row=row_max+1, column=colr).value = f'=COUNTIF({firstcell}:{passedcell},"<>"&"")'
        # total = sheet.cell(row=row_max+2, column=colr).coordinate

        firstcell_P = sheet.cell(row=ROW_PAD+1, column=colr).coordinate
        passedcell_P = sheet.cell(row=row_max, column=colr).coordinate

        sheet.cell(
            row=row_max+2, column=colr).value = f'=COUNTIFS({firstcell}:{passedcell},"REGULAR",{firstcell_P}:{passedcell_P},"passed")'
        passed = sheet.cell(row=row_max+2, column=colr).coordinate

        sheet.cell(row=row_max+2, column=colr -
                    1).value = "No of students Passed in Regular"

        sheet.cell(row=row_max+3, column=colr -
                    1).value = "Regular Pass Percentage"
        total = sheet.cell(row=row_max+1, column=colr).coordinate
        sheet.cell(
            row=row_max+3, column=colr).value = f"=ROUND({passed}/{total}*100,1)"

    
def fill_excel_total(exam, sheet, col, row, ROW_PAD):

    # print(col)
    Subject_Code = exam["Subject_Code"]
    Subject = exam

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['Mid1Marks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_Mid1Marks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['Mid2Marks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_Mid2Marks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['InternalMarks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_InternalMarks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['EndExamMarks']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_EndExamMarks"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = string_toFloat(f"{Subject['SubjectTotal']}")
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_SubjectTotal"
    col = col+1

    sheet.cell(row=row+1, column=col).value = f"{Subject['ExamMonthYear']}"
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_ExamMonthYear"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = f"{reg_sup(Subject['WholeOrSupply'])}"
    sheet.cell(
        row=ROW_PAD, column=col).value = f"{Subject_Code}_Writing_Regular_or_Suplly"
    col = col+1

    sheet.cell(
        row=row+1, column=col).value = f"{reg_stat(Subject['ExamStatus'])}"
    sheet.cell(row=ROW_PAD, column=col).value = f"{Subject_Code}_ExamStatus"

    sheet.cell(row=row+1+1, column=col-1).value = f" No of Students appeared"
    firstcell = sheet.cell(row=ROW_PAD+1, column=col-1).coordinate
    passedcell = sheet.cell(row=row+1, column=col-1).coordinate
    sheet.cell(
        row=row+1+1, column=col).value = f'=COUNTIF({firstcell}:{passedcell},"<>"&"")'
    total = sheet.cell(row=row+1+1, column=col).coordinate

    firstcell_P = sheet.cell(row=ROW_PAD+1, column=col).coordinate
    passedcell_P = sheet.cell(row=row+1, column=col).coordinate

    sheet.cell(
        row=row+1+2, column=col).value = f'=COUNTIFS({firstcell}:{passedcell},"REGULAR",{firstcell_P}:{passedcell_P},"PASSED")'
    passed = sheet.cell(row=row+1+2, column=col).coordinate

    sheet.cell(row=row+1+2, column=col -
               1).value = " No of students Passed in Regular"

    sheet.cell(row=row+1+3, column=col-1).value = "Regular Pass Percentage"
    sheet.cell(
        row=row+1+3, column=col).value = f"=ROUND({passed}/{total}*100,1)"

    col = col+1

    return col


def fill_excel_overview(details, sheet, row, ROW_PAD):
    col = 1
    for key in getList(details):
        if key != 'SEM_DETAILS':
            sheet.cell(
                row=row+1, column=col).value = string_toFloat(f'{details[key]}')
            col = col+1
        if key == 'SEM_DETAILS':
            for exam in details['SEM_DETAILS']:
                sheet.cell(
                    row=row+1, column=col).value = string_toFloat(f"{exam['Credits']}")
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_credits"
                col = col+1

                sheet.cell(
                    row=row+1, column=col).value = string_toFloat(f"{exam['SGPA']}")
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_SGPA"
                col = col+1

                sheet.cell(
                    row=row+1, column=col).value = f"{exam['PASSED_TYPE']}"
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_PASSED_TYPE"
                
                col = col+1

                sheet.cell(
                    row=row+1, column=col).value = f"{exam['PASSED_STATUS']}"
                sheet.cell(
                    row=ROW_PAD, column=col).value = f"{exam['Semester']}_PASSED_STATUS"
                
                
            

                

                col = col+1
                    


def getresult_Excel(pins):
    
    wb_obj = openpyxl.Workbook()
    sheet_over = wb_obj.create_sheet(index=0, title="Overview")
    sem_sheets = {}
    for i in range(6):
        sem_sheets[f"{i+1}SEM"] = wb_obj.create_sheet(
            index=i+1, title=f"{i+1}SEM")
        col = 0
        ROW_PAD = 3
        sem_sheets[f"{i+1}SEM"].cell(
            row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
        sem_sheets[f"{i+1}SEM"].cell(
            row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
        for trow in TOP_ROW_tot:
            sem_sheets[f"{i+1}SEM"].cell(row=ROW_PAD,
                                         column=col+1).value = f'{trow}'
            col = col+1

    col = 0
    ROW_PAD = 3
    sheet_over.cell(
        row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    sheet_over.cell(
        row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"

    for trow in TOP_ROW_gen:
        sheet_over.cell(row=ROW_PAD, column=col+1).value = f'{trow}'
        col = col+1

    for i in range(len(pins)):
        if __name__ == "__main__":
            print(f"geting result data for PIN: {pins[i]}....", end='')
        details, data = dip_res(pins[i])
        # print(details)
        fill_excel_overview(details, sheet_over, i+ROW_PAD, ROW_PAD)
        col = 0
        try:
            for key in getList(sem_sheets):
                col = 1
                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["Scheme"]}'
                col = col+1
                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["StudentName"]}'
                col = col+1
                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["Pin"]}'
                col = col+1

                sem_sheets[key].cell(
                    row=i+ROW_PAD+1, column=col).value = f'{data["Table"][0]["BranchCode"]}'
                col = col+1
                try:
                    for exam in data["Table2"]:
                        if key == exam['Semester']:
                           # print(col)

                            col = fill_excel_total(exam, sem_sheets[key], col, i+ROW_PAD, ROW_PAD)
                            #print("excuting")

                except:
                    sem_sheets[key].cell(
                        row=i+ROW_PAD+1, column=col).value = f'Details  not found'
                    col = col+1

        except:
            sem_sheets[key].cell(
                row=i+ROW_PAD+1, column=3).value = f'{pins[i]}'
            col = col+1

    
    row_max = sheet_over.max_row
    
    
    for YEARS in SEM_INSERT.keys():
        insert(sheet_over,YEARS,SEM_INSERT[YEARS],row_max,ROW_PAD)
    
    col_max = sheet_over.max_column + 1  
    last_insert(sheet_over,col_max,row_max,ROW_PAD)
    
    sheet_over.cell(row=row_max+1, column=len(TOP_ROW_gen)-1).value = f"No Diploma Completed Students:"
    firstcell_o = sheet_over.cell(row=ROW_PAD+1, column=len(TOP_ROW_gen)).coordinate
    passedcell_o = sheet_over.cell(row_max, column=len(TOP_ROW_gen)).coordinate
    sheet_over.cell(row=row_max+1, column=len(TOP_ROW_gen)).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"Completed_Diploma")'

    sheet_over.cell(row=row_max+1, column=col_max).value = f"No Students:"
    firstcell_o = sheet_over.cell(row=ROW_PAD+1, column=col_max).coordinate
    passedcell_o = sheet_over.cell(row_max, column=col_max).coordinate
    sheet_over.cell(row=row_max+1, column=col_max+1).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"<>"&"")'
    
    sheet_over.cell(row=row_max+1+1, column=col_max).value = f"No Diploma Completed without Backlogs Students:"
    firstcell_o = sheet_over.cell(row=ROW_PAD+1, column=col_max).coordinate
    passedcell_o = sheet_over.cell(row_max, column=col_max).coordinate
    sheet_over.cell(row=row_max+1+1, column=col_max+1).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"WITH_BACKLOG_PASSED_Diploma")'

    sheet_over.cell(row=row_max+1+1+1, column=col_max).value = f"No Diploma Completed with Backlogs Students:"
    firstcell_o = sheet_over.cell(row=ROW_PAD+1, column=col_max).coordinate
    passedcell_o = sheet_over.cell(row_max, column=col_max).coordinate
    sheet_over.cell(row=row_max+1+1+1, column=col_max+1).value = f'=COUNTIF({firstcell_o}:{passedcell_o},"WITH_BACKLOG_PASSED_Diploma")+COUNTIF({firstcell_o}:{passedcell_o},"WITHOUT_BACKLOG_PASSED_Diploma")'

    sheet_over.cell(row=row_max+1+1+1+1, column=col_max).value = f"Without BackLog Percentage"
    firstcell_o = sheet_over.cell(row=row_max+1, column=col_max+1).coordinate
    passedcell_o = sheet_over.cell(row=row_max+1+1, column=col_max+1).coordinate
    sheet_over.cell(row=row_max+1+1+1+1, column=col_max+1).value =f'=ROUND({passedcell_o}/{firstcell_o}*100,1)'

    sheet_over.cell(row=row_max+1+1+1+1+1, column=col_max).value = f"WithBackLog Percentage"
    firstcell_o = sheet_over.cell(row=row_max+1, column=col_max+1).coordinate
    passedcell_o = sheet_over.cell(row=row_max+1+1+1, column=col_max+1).coordinate
    sheet_over.cell(row=row_max+1+1+1+1+1, column=col_max+1).value =f'=ROUND({passedcell_o}/{firstcell_o}*100,1)'
    
    
    sheet_disc = wb_obj.create_sheet(index=0, title="Disclaimer")
    for YEARS in SEM_INSERT.keys():
        cal_per(sheet_over,SEM_INSERT[YEARS],row_max,ROW_PAD)
    
    sheet_disc.cell(
        row=1, column=1).value = "Result in excel by KEERTHI CHANDRA C,L/ECE, GMRPW Karimnagar "
    sheet_disc.cell(
        row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
    
    sheet_disc.cell(
        row=2, column=1).value = "for feedback Contact keerthichand.c@gmail.com"
    
    sheet_disc.merge_cells('A3:AB5') 
    sheet_disc.cell(
        row=3, column=1).value = '''
    
    


disclaimer:
    *These are not actual Results
    *Use this at your own discretion
    *for actual Results please visit TS-SBTET Website
    *OUTPUTS from this program are not related to TS-SBTET
    *This program is developed only to help ease work but not to misguide or harm in 
        any possible ways and with no ill intent
        '''
        

    
    return wb_obj


   
if __name__ == "__main__":
    coderintro()
    instructions_info()
    pins = getpins()
    wb_obj = getresult_Excel(pins)
    os.makedirs(f"{file_path}\\OUTPUT", exist_ok=True)
    wb_obj.save(f"{file_path}\\OUTPUT\\OUTPUT.xlsx")

    print(f'''
************************************************************************************

Finished and Saved in {file_path}\\OUTPUT\\OUTPUT.xlsx
check OUTPUTS folder for excel containing overview, 1SEM, 2SEM, 3SEM, 4SEM, 5SEM, 6SEM sheets.

************************************************************************************''')
      
    input("press ENTER to Exit......")
    os.system('cls')
    print()
    
